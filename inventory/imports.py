from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction

from audit.models import AuditLog
from products.models import Product, ProductCategory

from .models import HistoricalInventoryRecord, ImportJob, Supplier
from .services import InventoryService


TEMPLATES = {
    'products': ['sku', 'name', 'item_type', 'category', 'brand', 'model_number', 'buying_price', 'selling_price', 'reorder_threshold', 'serialized', 'tax_eligible'],
    'suppliers': ['company_name', 'contact_person', 'phone', 'email', 'physical_address', 'tin_vrn', 'notes', 'active'],
    'opening_stock': ['sku', 'quantity', 'serial_numbers', 'reason_notes'],
    'historical_purchases': ['date', 'reference', 'sku', 'description', 'quantity', 'unit_amount', 'notes'],
    'historical_sales': ['date', 'reference', 'sku', 'description', 'quantity', 'unit_amount', 'notes'],
}


def _openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise RuntimeError('Excel support requires the openpyxl dependency.') from exc
    return Workbook, load_workbook


def template_workbook(import_type: str) -> bytes:
    Workbook, _ = _openpyxl()
    if import_type not in TEMPLATES:
        raise ValueError('Unsupported import type.')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = import_type[:31]
    sheet.append(TEMPLATES[import_type])
    if import_type == 'products':
        sheet.append(['RTR-001', 'WiFi Router', 'physical', 'Routers', 'TP-Link', 'AX10', 100000, 150000, 2, 'no', 'yes'])
    elif import_type == 'opening_stock':
        sheet.append(['RTR-001', 10, '', 'Counted opening balance'])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def workbook_bytes(headers, rows, *, title='Export') -> bytes:
    Workbook, _ = _openpyxl()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31]
    sheet.append(list(headers))
    for row in rows:
        sheet.append(list(row))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _bool(value):
    return str(value or '').strip().lower() in {'1', 'true', 'yes', 'y', 'active'}


def _decimal(value, field, errors):
    try:
        return Decimal(str(value if value is not None else '')).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError):
        errors.append(f'{field} must be a number.')
        return Decimal('0.00')


def _date(value, errors):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value))
    except (TypeError, ValueError):
        errors.append('date must use YYYY-MM-DD.')
        return date.today()


def validate_workbook(*, organization, actor, import_type: str, uploaded_file) -> ImportJob:
    _, load_workbook = _openpyxl()
    if import_type not in TEMPLATES:
        raise ValueError('Unsupported import type.')
    errors = []
    rows = []
    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        headers = [str(value or '').strip() for value in next(values)]
    except Exception as exc:
        return ImportJob.objects.create(
            organization=organization, tenant=organization, import_type=import_type,
            status=ImportJob.Status.FAILED, file_name=getattr(uploaded_file, 'name', 'upload.xlsx'),
            row_count=0, error_count=1, errors=[{'row': 0, 'errors': [f'Workbook could not be read: {exc}']}], created_by=actor,
        )
    required = TEMPLATES[import_type]
    missing = [header for header in required if header not in headers]
    if missing:
        errors.append({'row': 1, 'errors': [f"Missing columns: {', '.join(missing)}"]})
    else:
        positions = {header: headers.index(header) for header in required}
        for number, values_row in enumerate(values, start=2):
            if not any(value not in (None, '') for value in values_row):
                continue
            raw = {header: values_row[index] if index < len(values_row) else None for header, index in positions.items()}
            row_errors = []
            normalized = _validate_row(organization=organization, import_type=import_type, raw=raw, errors=row_errors)
            if row_errors:
                errors.append({'row': number, 'errors': row_errors})
            normalized['_row'] = number
            rows.append(normalized)
    return ImportJob.objects.create(
        organization=organization,
        tenant=organization,
        import_type=import_type,
        status=ImportJob.Status.VALIDATED if not errors else ImportJob.Status.FAILED,
        file_name=getattr(uploaded_file, 'name', 'upload.xlsx'),
        affects_live_stock=import_type == 'opening_stock',
        row_count=len(rows),
        error_count=len(errors),
        errors=errors,
        validated_rows=rows,
        created_by=actor,
    )


def _validate_row(*, organization, import_type, raw, errors):
    if import_type == 'products':
        sku = str(raw['sku'] or '').strip().upper()
        name = str(raw['name'] or '').strip()
        item_type = str(raw['item_type'] or 'physical').strip().lower()
        if not sku:
            errors.append('sku is required.')
        if not name:
            errors.append('name is required.')
        if item_type not in Product.ItemType.values:
            errors.append('item_type must be physical or service.')
        if Product.objects.unscoped().filter(tenant=organization, sku__iexact=sku).exists():
            errors.append('sku already exists.')
        buying = _decimal(raw['buying_price'], 'buying_price', errors)
        selling = _decimal(raw['selling_price'], 'selling_price', errors)
        threshold = _decimal(raw['reorder_threshold'], 'reorder_threshold', errors)
        if min(buying, selling, threshold) < 0:
            errors.append('prices and threshold cannot be negative.')
        return {
            'sku': sku, 'name': name, 'item_type': item_type, 'category': str(raw['category'] or '').strip(),
            'brand': str(raw['brand'] or '').strip(), 'model_number': str(raw['model_number'] or '').strip(),
            'buying_price': str(buying), 'selling_price': str(selling), 'reorder_threshold': str(threshold),
            'serialized': _bool(raw['serialized']), 'tax_eligible': _bool(raw['tax_eligible']),
        }
    if import_type == 'suppliers':
        company = str(raw['company_name'] or '').strip()
        if not company:
            errors.append('company_name is required.')
        if Supplier.objects.unscoped().filter(tenant=organization, company_name__iexact=company).exists():
            errors.append('supplier already exists.')
        return {key: str(raw[key] or '').strip() for key in TEMPLATES['suppliers']} | {'active': _bool(raw['active'])}
    if import_type == 'opening_stock':
        sku = str(raw['sku'] or '').strip().upper()
        product = Product.objects.unscoped().filter(tenant=organization, sku__iexact=sku).first()
        if product is None:
            errors.append('sku was not found in this tenant.')
        elif not product.track_stock or product.item_type == Product.ItemType.SERVICE:
            errors.append('sku is not stock tracked.')
        quantity = _decimal(raw['quantity'], 'quantity', errors)
        if quantity <= 0:
            errors.append('quantity must be greater than zero.')
        serials = [item.strip().upper() for item in str(raw['serial_numbers'] or '').replace(',', '\n').splitlines() if item.strip()]
        if product and product.is_serialized and len(serials) != int(quantity):
            errors.append('serialized opening stock needs one serial per unit.')
        return {'sku': sku, 'quantity': str(quantity), 'serial_numbers': serials, 'reason_notes': str(raw['reason_notes'] or '').strip()}
    record_date = _date(raw['date'], errors)
    quantity = _decimal(raw['quantity'], 'quantity', errors)
    unit_amount = _decimal(raw['unit_amount'], 'unit_amount', errors)
    if quantity <= 0 or unit_amount < 0:
        errors.append('quantity must be positive and unit_amount cannot be negative.')
    description = str(raw['description'] or '').strip()
    if not description:
        errors.append('description is required.')
    return {
        'date': record_date.isoformat(), 'reference': str(raw['reference'] or '').strip(),
        'sku': str(raw['sku'] or '').strip().upper(), 'description': description,
        'quantity': str(quantity), 'unit_amount': str(unit_amount), 'notes': str(raw['notes'] or '').strip(),
    }


@transaction.atomic
def commit_import(*, organization, actor, job_id: int):
    job = ImportJob.objects.unscoped().select_for_update().filter(pk=job_id, tenant=organization).first()
    if job is None:
        raise ValueError('Import job not found.')
    if job.status != ImportJob.Status.VALIDATED or job.error_count:
        raise ValueError('Only a validation with zero errors can be committed.')
    for row in job.validated_rows:
        if job.import_type == 'products':
            category = None
            if row['category']:
                category, _ = ProductCategory.objects.unscoped().get_or_create(
                    tenant=organization, name=row['category'],
                    defaults={'organization': organization, 'description': '', 'is_active': True},
                )
            Product.objects.create(
                organization=organization, tenant=organization, sku=row['sku'], name=row['name'],
                item_type=row['item_type'], catalog_category=category, brand=row['brand'], model_number=row['model_number'],
                buying_price=Decimal(row['buying_price']), selling_price=Decimal(row['selling_price']),
                retail_price=Decimal(row['selling_price']), reorder_threshold=Decimal(row['reorder_threshold']),
                is_serialized=row['serialized'], track_stock=row['item_type'] == Product.ItemType.PHYSICAL,
                tax_eligible=row['tax_eligible'], quantity=Decimal('0.00'), stock=0, measure_unit='Unit',
            )
        elif job.import_type == 'suppliers':
            Supplier.objects.create(
                organization=organization, tenant=organization, company_name=row['company_name'],
                contact_person=row['contact_person'], phone=row['phone'], email=row['email'],
                physical_address=row['physical_address'], tin_vrn=row['tin_vrn'], notes=row['notes'],
                is_active=row['active'], created_by=actor,
            )
        elif job.import_type == 'opening_stock':
            product = Product.objects.unscoped().get(tenant=organization, sku=row['sku'])
            InventoryService.adjust_stock(
                organization=organization, product_id=product.pk, quantity_delta=Decimal(row['quantity']),
                reason='opening_balance', notes=row['reason_notes'], serial_numbers=row['serial_numbers'], actor=actor,
            )
        else:
            record_type = HistoricalInventoryRecord.RecordType.PURCHASE if job.import_type == 'historical_purchases' else HistoricalInventoryRecord.RecordType.SALE
            HistoricalInventoryRecord.objects.create(
                organization=organization, tenant=organization, record_type=record_type,
                record_date=date.fromisoformat(row['date']), reference=row['reference'], sku=row['sku'],
                description=row['description'], quantity=Decimal(row['quantity']), unit_amount=Decimal(row['unit_amount']),
                total_amount=Decimal(row['quantity']) * Decimal(row['unit_amount']), notes=row['notes'], import_job=job,
            )
    job.status = ImportJob.Status.COMMITTED
    job.save(update_fields=['status'])
    AuditLog.objects.create(
        organization=organization, tenant=organization, actor=actor, performed_by=actor,
        action='inventory.import.committed', action_type='inventory.import.committed', object_type='ImportJob',
        object_id=str(job.pk), document_id=str(job.pk),
        metadata={'import_type': job.import_type, 'row_count': job.row_count, 'affects_live_stock': job.affects_live_stock},
    )
    return job
